from yahoo_fin.stock_info import *
import openpyxl as xl

wb = xl.load_workbook('Orders.xlsx')
sheet = wb['TEST']
for row in sheet['A2:H70']:  # 70 seems like the right cut off
    for cell in row:
        if cell.value == 'BYND':
            print(f'In order: R{cell.row} C{cell.column}')


def update_stocks():
    try:
        print("Processing...")
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)

            if cell.value is not None:  # lol this line i so funny to me
                # print(cell.value)
                ticker = cell.value
                price = round(get_live_price(ticker), 2)
                new_price_cell = sheet.cell(row, 7)
                new_price_cell.value = price
                # print(new_price_cell.value)
                share_amount_cell = sheet.cell(row, 4)
                current_value = share_amount_cell.value * new_price_cell.value
                current_value_cell = sheet.cell(row, 8)
                current_value_cell.value = current_value
                wb.save('Orders.xlsx')
                print("Update Complete")
    except PermissionError:
        print('Please Exit the workbook')

def print_sheets():
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value is not None :
                print(f'{cell.value}-', end="")
        print()

