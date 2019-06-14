class Order:
    def __init__(self, order_name, ticker, buy_price, stock_amount, value_atpurchase, date):
        self.order_name = order_name
        self.ticker = ticker
        self.buy_price = buy_price
        self.stock_amount = stock_amount
        self.value_atpurchase = value_atpurchase
        self.date = date

    def save_order(self):
        import openpyxl as xl
        wb = xl.load_workbook('Orders.xlsx')
        sheet = wb['TEST']
        freerow = sheet.max_row + 1
        order_list = [self.order_name, self.ticker, self.buy_price, self.stock_amount, self.value_atpurchase, self.date]
        for col, val in enumerate(order_list, start=1):
            sheet.cell(row=freerow, column=col).value = val
        wb.save('Orders.xlsx')
        wb.close()


