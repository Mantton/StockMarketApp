import openpyxl as xl
from yahoo_fin.stock_info import *

wb = xl.load_workbook('Orders.xlsx')
sheet = wb['TEST']

wb2 = xl.load_workbook('Data.xlsx')
sheet2 = wb2['Sheet1']




def update_stocks():  # As the name implies this function updates the data of the 'orders.xlsx' file
    try:
        print("Processing...")
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)

            if cell.value is not None:  # lol this line i so funny to me
                # print(cell.value)
                ticker = cell.value
                price = round(get_live_price(ticker), 2)
                new_price_cell = sheet.cell(row, 7)
                new_price_cell.value = price
                # print(new_price_cell.value)
                share_amount_cell = sheet.cell(row, 4)
                current_value = share_amount_cell.value * new_price_cell.value
                current_value_cell = sheet.cell(row, 8)
                current_value_cell.value = round(current_value, 2)
        wb.save('Orders.xlsx')
        print("50% COMPLETE")
        wb.close()
        print("ORDER DATABASE UPDATE COMPLETE")
        print("RUNNING DATASHEET UPDATES\n")
    except PermissionError:
        print('Please Exit the workbook')


def clear_orders():
    for row in sheet['A2:H70']:  # 70 seems like the right cut off
        for cell in row:
            if cell.value is not None:
                cell.value = None
    wb.save('Orders.xlsx')
    for row in sheet2['A2:H70']:  # 70 seems like the right cut off
        for cell in row:
            if cell.value is not None:
                cell.value = None
    wb2.save('Data.xlsx')


def print_sheets():
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value is not None:
                print(f' {cell.value}  -', end="  ")
        print()


def print_datasheets():
    for row in sheet2.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value is not None:
                print(f' {cell.value}  -', end="  ")
        print()


def update_datasheet():
    print('Updating Data sheet...')

    ticker_list = []

    for row in sheet['B2:B70']:
        for cell in row:
            if cell.value is not None:
                ticker_list.append(cell.value)

    ticker_list = list(dict.fromkeys(ticker_list))

    data_list = []

    for row in range(2, sheet2.max_row + 1):  # 70 seems like the right cut off
        # for cell in row:
        cell = sheet2.cell(row=row, column=2)
        if cell.value is not None:
            data_list.append(cell.value)

    newdata_list = []
    for item in ticker_list:
        if item not in data_list:
            newdata_list.append(item)

    for ticker in ticker_list:
        total_stock = 0
        freerow = sheet2.max_row + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)

            if cell.value == ticker:
                stock_cell = sheet.cell(row, 4)
                stock = stock_cell.value
                total_stock += stock

        test = [ticker, total_stock]

        for row in range(1, freerow):  # 70 seems like the right cut off
            # for cell in row:
            cell = sheet2.cell(row=row, column=2)

            if cell.value == ticker:
                stock_cell = sheet2.cell(row=row, column=3)
                stock_cell.value = total_stock
                cell.value = ticker

                break

    for data in newdata_list:
        total_stock = 0
        freerow = sheet2.max_row + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)

            if cell.value == data:
                stock_cell = sheet.cell(row, 4)
                stock = stock_cell.value
                total_stock += stock
        array = [data, total_stock]

        for col, val in enumerate(array, start=2):
            sheet2.cell(row=freerow, column=col).value = val
        wb2.save('Data.xlsx')

    print('Loading...')
    for row in range(2, sheet2.max_row + 1):
        cell = sheet2.cell(row, 2)

        if cell.value is not None:  # lol this line i so funny to me
            # print(cell.value)
            ticker = cell.value
            price = round(get_live_price(ticker), 2)
            new_price_cell = sheet2.cell(row, 4)
            new_price_cell.value = price
            # print(new_price_cell.value)
            current_value_cell = sheet2.cell(row, 5)
            share_amount_cell = sheet2.cell(row, 3)
            old_value_cell = sheet2.cell(row, 6)
            old_value_cell.value = current_value_cell.value

            current_value = share_amount_cell.value * new_price_cell.value
            # current_value_cell = sheet2.cell(row, 5)
            current_value_cell.value = round(current_value, 2)
            diff = current_value - old_value_cell.value
            per_diff = ((current_value - old_value_cell.value) / old_value_cell.value) * 100
            sheet2.cell(row, 7).value = round(per_diff, 2)
            sheet2.cell(row, 8).value = round(diff, 2)
            print(f'Processing...{row}/{sheet2.max_row - 1}')
    print('UPDATE COMPLETE')
    wb2.save('Data.xlsx')


def total_invested():
    current_total = 0
    last_total = 0
    diff = 0
    t_diff = 0
    for row in range(2, sheet2.max_row + 1):
        new_total_cell = sheet2.cell(row, 5)
        old_total_cell = sheet2.cell(row, 6)
        current_total += new_total_cell.value
        last_total += old_total_cell.value
        diff = ((current_total - last_total) / last_total) * 100
        t_diff = current_total - last_total
        t_diff = round(t_diff, 1)
    print(f'NEW TOTAL INVESTED : ${round(current_total, 1)}\n'
          f'OLD TOTAL INVESTED : ${round(last_total, 1)}\n'
          f'MONETARY GAIN / LOSS : ${t_diff}\n'
          f'PERCENT DIFF: {round(diff, 1)}%')


def sell_stocks(ticker, amount):
    # unable to import the backend script so im just gonna create a parameter for it
    ticker_list = []

    for row in range(2, sheet.max_row + 1):
        cell = sheet2.cell(row, 2)
        if cell.value is not None:
            # print(cell.value)
            ticker_list.append(cell.value)

    ticker_list = list(dict.fromkeys(ticker_list))
    # print(ticker_list)

    if ticker in ticker_list:
        for row in range(2, sheet.max_row + 1):
            ticker_cell = sheet.cell(row, 2)
            amount_cell = sheet.cell(row, 4)

            if ticker_cell.value == ticker:
                if amount_cell.value > amount:
                    amount_cell.value -= amount

                    # Price scraping
                    price = round(get_live_price(ticker))
                    balance = price * amount
                    print('Success')
                    return balance
                else:
                    return None










if __name__ == '__main__':
    total_invested()
    wb2.close()
    wb.close()
